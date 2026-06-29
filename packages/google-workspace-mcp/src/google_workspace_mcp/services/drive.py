"""
Google Drive service implementation for file operations.
Provides comprehensive file management capabilities through Google Drive API.
"""

import base64
import binascii
import csv
import io
import logging
import mimetypes
from typing import Any

from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

from google_workspace_mcp.services.base import BaseGoogleService

logger = logging.getLogger(__name__)

# Office Open XML mime types that are uploaded as opaque binaries to Drive
# (i.e. NOT native Google Workspace files). Without dedicated handling these
# fall through to a base64 dump that no downstream tool can read.
XLSX_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
NATIVE_SHEET_MIME_TYPE = "application/vnd.google-apps.spreadsheet"

# Native Google Workspace types are not byte-downloadable; a download_url for
# them would be meaningless, so we skip building one for these.
_GOOGLE_NATIVE_MIME_PREFIX = "application/vnd.google-apps."


def _direct_download_url(file_id: str) -> str:
    """Build a non-redirecting direct-download URL for a Drive file id.

    Drive's ``webContentLink`` (``drive.google.com/uc?...&export=download``)
    303-redirects to a virus-scan confirmation page for large files (e.g.
    videos), which breaks consumers that don't follow redirects. The
    ``drive.usercontent.google.com/download`` host with ``confirm=t`` serves
    the bytes directly with no redirect and no scan interstitial.
    """
    return (
        f"https://drive.usercontent.google.com/download?id={file_id}"
        "&export=download&confirm=t"
    )


def _with_download_url(file_meta: dict[str, Any]) -> dict[str, Any]:
    """Add a ``download_url`` to a file-metadata dict when applicable.

    Adds the non-redirecting direct-download URL (see ``_direct_download_url``)
    for byte-downloadable files that have an ``id``. Native Google Workspace
    files (Docs/Sheets/Slides) are skipped since they aren't byte-downloadable;
    they get ``download_url=None`` so callers can rely on the key existing.
    Mutates and returns the same dict.
    """
    if not isinstance(file_meta, dict):
        return file_meta
    file_id = file_meta.get("id")
    mime = file_meta.get("mimeType") or ""
    if file_id and not mime.startswith(_GOOGLE_NATIVE_MIME_PREFIX):
        file_meta["download_url"] = _direct_download_url(file_id)
    else:
        file_meta.setdefault("download_url", None)
    return file_meta


class DriveService(BaseGoogleService):
    """
    Service for interacting with Google Drive API.
    """

    def __init__(self):
        """Initialize the Drive service."""
        super().__init__("drive", "v3")

    def search_files(
        self,
        query: str,
        page_size: int = 10,
        shared_drive_id: str | None = None,
        include_shared_drives: bool = False,
    ) -> list[dict[str, Any]]:
        """
        Search for files in Google Drive.

        Args:
            query: Search query string
            page_size: Maximum number of files to return (1-1000)
            shared_drive_id: Optional shared drive ID to search within a specific shared drive
            include_shared_drives: When True (and no specific shared_drive_id is
                given), search across My Drive AND all shared drives/folders the
                user can access (corpora="allDrives"). Needed to find files that
                live in a Shared Drive by name. Defaults False to preserve the
                user-only default behavior.

        Returns:
            List of file metadata dictionaries (id, name, mimeType, etc.) or an error dictionary
        """
        try:
            logger.info(
                f"Searching files with query: '{query}', page_size: {page_size}, "
                f"shared_drive_id: {shared_drive_id}, include_shared_drives: {include_shared_drives}"
            )

            # page_size is the desired TOTAL number of results. The Drive API
            # returns at most 1000 per request, so request in chunks and follow
            # nextPageToken until we have enough (previously results past the
            # first page were silently lost).
            desired_total = max(1, page_size)

            # Build list parameters with shared drive support. `parents` is
            # included so folder-aware callers can inspect file location.
            # `webContentLink` is the direct-download URL (present for binary
            # files shared "Anyone with the link"); callers that move bytes
            # server-side — e.g. a media-upload pipeline chaining the link into
            # another service — need it so the file never has to be downloaded
            # through the agent. `webViewLink` is the human viewer page, not a
            # fetchable asset.
            list_params = {
                "q": query,  # Use the query directly without modification
                "fields": "nextPageToken, files(id, name, mimeType, modifiedTime, size, webViewLink, webContentLink, iconLink, parents)",
                "supportsAllDrives": True,
                "includeItemsFromAllDrives": True,
            }

            if shared_drive_id:
                list_params["driveId"] = shared_drive_id
                list_params["corpora"] = (
                    "drive"  # Search within the specified shared drive
                )
            elif include_shared_drives:
                # Search My Drive + all accessible shared drives/folders.
                list_params["corpora"] = "allDrives"
            else:
                list_params["corpora"] = (
                    "user"  # Default to user's files if no specific shared drive ID
                )

            files: list[dict[str, Any]] = []
            page_token: str | None = None
            while len(files) < desired_total:
                list_params["pageSize"] = min(desired_total - len(files), 1000)
                if page_token:
                    list_params["pageToken"] = page_token
                results = self.service.files().list(**list_params).execute()
                files.extend(results.get("files", []))
                page_token = results.get("nextPageToken")
                if not page_token:
                    break

            files = files[:desired_total]
            for f in files:
                _with_download_url(f)
            logger.info(f"Found {len(files)} files matching query '{query}'")
            return files

        except Exception as e:
            return self.handle_api_error("search_files", e)

    def read_file_content(self, file_id: str) -> dict[str, Any] | None:
        """
        Read the content of a file from Google Drive.

        Args:
            file_id: The ID of the file to read

        Returns:
            Dict containing mimeType and content (possibly base64 encoded)
        """
        try:
            # Get file metadata (supportsAllDrives so Shared Drive files resolve)
            file_metadata = (
                self.service.files()
                .get(fileId=file_id, fields="mimeType, name", supportsAllDrives=True)
                .execute()
            )

            original_mime_type = file_metadata.get("mimeType")
            file_name = file_metadata.get("name", "Unknown")

            logger.info(
                f"Reading file '{file_name}' ({file_id}) with mimeType: {original_mime_type}"
            )

            # Handle Google Workspace files by exporting
            if original_mime_type.startswith("application/vnd.google-apps."):
                return self._export_google_file(file_id, file_name, original_mime_type)
            return self._download_regular_file(file_id, file_name, original_mime_type)

        except Exception as e:
            return self.handle_api_error("read_file", e)

    def get_file_metadata(self, file_id: str) -> dict[str, Any]:
        """
        Get metadata information for a file from Google Drive.

        Args:
            file_id: The ID of the file to get metadata for

        Returns:
            Dict containing file metadata or error information
        """
        try:
            if not file_id:
                return {"error": True, "message": "File ID cannot be empty"}

            logger.info(f"Getting metadata for file with ID: {file_id}")

            # Retrieve file metadata with comprehensive field selection
            file_metadata = (
                self.service.files()
                .get(
                    fileId=file_id,
                    fields="id, name, mimeType, size, createdTime, modifiedTime, "
                    "webViewLink, webContentLink, iconLink, parents, owners, "
                    "shared, trashed, capabilities, permissions, "
                    "description, starred, explicitlyTrashed",
                    supportsAllDrives=True,
                )
                .execute()
            )

            logger.info(
                f"Successfully retrieved metadata for file: {file_metadata.get('name', 'Unknown')}"
            )
            return file_metadata

        except Exception as e:
            return self.handle_api_error("get_file_metadata", e)

    def create_folder(
        self,
        folder_name: str,
        parent_folder_id: str | None = None,
        shared_drive_id: str | None = None,
    ) -> dict[str, Any]:
        """
        Create a new folder in Google Drive.

        Args:
            folder_name: The name for the new folder
            parent_folder_id: Optional parent folder ID to create the folder within
            shared_drive_id: Optional shared drive ID to create the folder in a shared drive

        Returns:
            Dict containing the created folder information or error details
        """
        try:
            if not folder_name or not folder_name.strip():
                return {"error": True, "message": "Folder name cannot be empty"}

            logger.info(
                f"Creating folder '{folder_name}' with parent_folder_id: {parent_folder_id}, shared_drive_id: {shared_drive_id}"
            )

            # Build folder metadata
            folder_metadata = {
                "name": folder_name.strip(),
                "mimeType": "application/vnd.google-apps.folder",
            }

            # Set parent folder if specified
            if parent_folder_id:
                folder_metadata["parents"] = [parent_folder_id]
            elif shared_drive_id:
                # If shared drive is specified but no parent, set shared drive as parent
                folder_metadata["parents"] = [shared_drive_id]

            # Create the folder with shared drive support
            create_params = {
                "body": folder_metadata,
                "fields": "id, name, parents, webViewLink, createdTime",
                "supportsAllDrives": True,
            }

            if shared_drive_id:
                create_params["driveId"] = shared_drive_id

            created_folder = self.service.files().create(**create_params).execute()

            logger.info(
                f"Successfully created folder '{folder_name}' with ID: {created_folder.get('id')}"
            )
            return created_folder

        except Exception as e:
            return self.handle_api_error("create_folder", e)

    def _export_google_file(
        self, file_id: str, file_name: str, mime_type: str
    ) -> dict[str, Any]:
        """Export a Google Workspace file in an appropriate format."""
        # Determine export format
        export_mime_type = None
        if mime_type == "application/vnd.google-apps.document":
            export_mime_type = "text/markdown"  # Consistently use markdown for docs
        elif mime_type == "application/vnd.google-apps.spreadsheet":
            export_mime_type = "text/csv"
        elif mime_type == "application/vnd.google-apps.presentation":
            export_mime_type = "text/plain"
        elif mime_type == "application/vnd.google-apps.drawing":
            export_mime_type = "image/png"

        if not export_mime_type:
            logger.warning(f"Unsupported Google Workspace type: {mime_type}")
            return {
                "error": True,
                "error_type": "unsupported_type",
                "message": f"Unsupported Google Workspace file type: {mime_type}",
                "mimeType": mime_type,
                "operation": "_export_google_file",
            }

        # Export the file
        try:
            request = self.service.files().export_media(
                fileId=file_id, mimeType=export_mime_type
            )

            content_bytes = self._download_content(request)
            if isinstance(content_bytes, dict) and content_bytes.get("error"):
                return content_bytes

            # Process the content based on MIME type
            if export_mime_type.startswith("text/"):
                try:
                    content = content_bytes.decode("utf-8")
                    return {
                        "mimeType": export_mime_type,
                        "content": content,
                        "encoding": "utf-8",
                    }
                except UnicodeDecodeError:
                    content = base64.b64encode(content_bytes).decode("utf-8")
                    return {
                        "mimeType": export_mime_type,
                        "content": content,
                        "encoding": "base64",
                    }
            else:
                content = base64.b64encode(content_bytes).decode("utf-8")
                return {
                    "mimeType": export_mime_type,
                    "content": content,
                    "encoding": "base64",
                }
        except Exception as e:
            return self.handle_api_error("_export_google_file", e)

    def _download_regular_file(
        self, file_id: str, file_name: str, mime_type: str
    ) -> dict[str, Any]:
        """Download a regular (non-Google Workspace) file."""
        request = self.service.files().get_media(fileId=file_id, supportsAllDrives=True)

        content_bytes = self._download_content(request)
        if isinstance(content_bytes, dict) and content_bytes.get("error"):
            return content_bytes

        # Process text files
        if mime_type.startswith("text/") or mime_type == "application/json":
            try:
                content = content_bytes.decode("utf-8")
                return {"mimeType": mime_type, "content": content, "encoding": "utf-8"}
            except UnicodeDecodeError:
                logger.warning(
                    f"UTF-8 decoding failed for file {file_id} ('{file_name}', {mime_type}). Using base64."
                )
                content = base64.b64encode(content_bytes).decode("utf-8")
                return {
                    "mimeType": mime_type,
                    "content": content,
                    "encoding": "base64",
                }

        # Uploaded (non-native) .xlsx workbooks: parse into CSV-per-sheet text so
        # the content is readable. Without this, an uploaded .xlsx is neither a
        # native Google Sheet (no export path) nor text, so it would fall through
        # to an unusable base64 dump.
        if mime_type == XLSX_MIME_TYPE:
            extracted = self._extract_xlsx_text(content_bytes, file_name)
            if extracted is not None:
                return {
                    "mimeType": "text/csv",
                    "content": extracted,
                    "encoding": "utf-8",
                    "sourceMimeType": mime_type,
                }
            logger.warning(
                f"xlsx extraction failed for '{file_name}' ({file_id}); returning base64."
            )

        # Binary file
        content = base64.b64encode(content_bytes).decode("utf-8")
        return {"mimeType": mime_type, "content": content, "encoding": "base64"}

    @staticmethod
    def _extract_xlsx_text(content_bytes: bytes, file_name: str) -> str | None:
        """Render an .xlsx workbook as CSV text, one block per sheet.

        Returns None if the workbook cannot be parsed, so the caller can fall
        back to base64. Multi-sheet workbooks are separated by a header line and
        a blank line; each sheet is emitted as standard CSV preserving rows and
        columns.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed; cannot extract .xlsx text.")
            return None

        try:
            workbook = load_workbook(
                io.BytesIO(content_bytes), read_only=True, data_only=True
            )
        except Exception as e:
            logger.warning(f"Failed to open .xlsx '{file_name}': {e}")
            return None

        try:
            blocks: list[str] = []
            multi_sheet = len(workbook.sheetnames) > 1
            for sheet in workbook.worksheets:
                buffer = io.StringIO()
                writer = csv.writer(buffer, lineterminator="\n")
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(
                        ["" if cell is None else cell for cell in row]
                    )
                sheet_csv = buffer.getvalue().rstrip("\r\n")
                if multi_sheet:
                    blocks.append(f"# Sheet: {sheet.title}\n{sheet_csv}")
                else:
                    blocks.append(sheet_csv)
            return "\n\n".join(blocks)
        finally:
            workbook.close()

    def _download_content(self, request) -> bytes:
        """Download content from a request."""
        try:
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)

            done = False
            while not done:
                status, done = downloader.next_chunk()

            return fh.getvalue()

        except Exception as e:
            return self.handle_api_error("download_content", e)

    def upload_file_content(
        self,
        filename: str,
        content_base64: str,
        parent_folder_id: str | None = None,
        shared_drive_id: str | None = None,
        share: bool = False,
    ) -> dict[str, Any]:
        """
        Upload a file to Google Drive using its content.

        Args:
            filename: The name for the file in Google Drive.
            content_base64: Base64 encoded content of the file.
            parent_folder_id: Optional parent folder ID.
            shared_drive_id: Optional shared drive ID.
            share: When True, grant an "anyone with the link → reader" permission
                so the file is fetchable by its webContentLink without
                authentication. Defaults to False (the file stays private);
                opt in when a caller needs an unauthenticated download URL.

        Returns:
            Dict containing file metadata (including webContentLink) on success,
            or error information on failure.
        """
        try:
            logger.info(f"Uploading file '{filename}' from content.")

            # Decode the base64 content
            try:
                content_bytes = base64.b64decode(content_base64, validate=True)
            except (ValueError, TypeError, binascii.Error) as e:
                logger.error(f"Invalid base64 content for file '{filename}': {e}")
                return {
                    "error": True,
                    "error_type": "invalid_content",
                    "message": "Invalid base64 encoded content provided.",
                    "operation": "upload_file_content",
                }

            # An empty (or whitespace-only) base64 string decodes to zero bytes,
            # which would silently upload a 0-byte file. Reject it explicitly —
            # this also catches the case where a large blob arrived empty as a
            # tool argument.
            if not content_bytes:
                logger.error(f"Decoded content for file '{filename}' is empty.")
                return {
                    "error": True,
                    "error_type": "invalid_content",
                    "message": (
                        "Decoded file content is empty (0 bytes). The content "
                        "may have been lost in transit; provide non-empty bytes."
                    ),
                    "operation": "upload_file_content",
                }

            # Get file MIME type from filename
            mime_type, _ = mimetypes.guess_type(filename)
            if mime_type is None:
                mime_type = "application/octet-stream"

            file_metadata = {"name": filename}
            if parent_folder_id:
                file_metadata["parents"] = [parent_folder_id]
            elif shared_drive_id:
                file_metadata["parents"] = [shared_drive_id]

            # Use MediaIoBaseUpload for in-memory content
            media = MediaIoBaseUpload(io.BytesIO(content_bytes), mimetype=mime_type)

            create_params = {
                "body": file_metadata,
                "media_body": media,
                "fields": "id,name,mimeType,modifiedTime,size,webViewLink,webContentLink",
                "supportsAllDrives": True,
            }
            if shared_drive_id:
                create_params["driveId"] = shared_drive_id

            file = self.service.files().create(**create_params).execute()
            file_id = file.get("id")
            logger.info(f"Successfully uploaded file with ID: {file_id}")

            if share and file_id:
                # The upload already succeeded; from here we only enrich the
                # result (grant sharing, refresh the link). Failures must NOT
                # discard the uploaded file, so the grant and the re-fetch are
                # guarded independently and never re-raise.
                share_ok = self._grant_anyone_reader(file_id, shared_drive_id)
                file["shared"] = share_ok["shared"]
                if not share_ok["shared"]:
                    file["share_error"] = share_ok["share_error"]
                elif "webContentLink" not in file or not file.get("webContentLink"):
                    # webContentLink is typically only populated after sharing,
                    # so re-fetch to return an accurate direct-download URL.
                    refreshed = self._refetch_metadata(file_id)
                    if refreshed is not None:
                        file = {**refreshed, "shared": True}

            # Guarantee the key exists so callers can rely on .get / indexing
            # (webContentLink is absent for Google-native files and some
            # configurations even when shared).
            file.setdefault("webContentLink", None)
            # Add the non-redirecting direct-download URL (reliable for large
            # binary files, unlike webContentLink which 303s).
            _with_download_url(file)
            return file

        except HttpError as e:
            return self.handle_api_error("upload_file_content", e)
        except Exception as e:
            logger.error(f"Non-API error in upload_file_content: {str(e)}")
            return {
                "error": True,
                "error_type": "local_error",
                "message": f"Error uploading file from content: {str(e)}",
                "operation": "upload_file_content",
            }

    def _grant_anyone_reader(
        self, file_id: str, shared_drive_id: str | None = None
    ) -> dict[str, Any]:
        """Grant "anyone with the link → reader" on a file. Never raises.

        Returns {"shared": True} on success, or {"shared": False,
        "share_error": <message>} on failure. A 403 typically means an
        organization/Shared Drive policy forbids "anyone with the link"
        sharing, which is reported distinctly from other failures.
        """
        try:
            self.service.permissions().create(
                fileId=file_id,
                body={"type": "anyone", "role": "reader"},
                supportsAllDrives=True,
            ).execute()
            return {"shared": True}
        except HttpError as e:
            status = e.resp.status if e.resp else None
            if status == 403:
                msg = (
                    "Anyone-with-link sharing is not permitted for this file "
                    "(likely an organization or Shared Drive sharing policy)."
                )
            else:
                msg = f"Failed to grant anyone-with-link access (HTTP {status})."
            logger.warning(f"Sharing {file_id} failed: {e}")
            return {"shared": False, "share_error": msg}
        except Exception as e:
            logger.warning(f"Sharing {file_id} failed (non-API): {e}")
            return {
                "shared": False,
                "share_error": "Failed to grant anyone-with-link access.",
            }

    def _refetch_metadata(self, file_id: str) -> dict[str, Any] | None:
        """Re-read a file's metadata (incl. webContentLink). None on failure.

        Used after sharing to pick up the now-populated webContentLink. A
        failure here is non-fatal — the upload already succeeded — so it returns
        None instead of raising.
        """
        try:
            return (
                self.service.files()
                .get(
                    fileId=file_id,
                    fields="id,name,mimeType,modifiedTime,size,webViewLink,webContentLink",
                    supportsAllDrives=True,
                )
                .execute()
            )
        except Exception as e:
            logger.warning(f"Could not re-fetch metadata for {file_id}: {e}")
            return None

    def convert_xlsx_to_google_sheet(
        self,
        source_file_id: str,
        new_name: str | None = None,
        parent_folder_id: str | None = None,
    ) -> dict[str, Any]:
        """
        Convert an .xlsx file already in Drive into a native Google Sheet.

        Drive performs the conversion server-side when a file is copied with a
        Google Workspace target mimeType, so this creates a NEW native Sheet and
        leaves the original .xlsx untouched. The resulting Sheet is fully
        editable by all Sheets tools (ranges, formulas, formatting), unlike an
        uploaded .xlsx which can only be read.

        Args:
            source_file_id: ID of the .xlsx file in Drive to convert.
            new_name: Optional name for the new Sheet. Defaults to the source
                file's name (with any .xlsx extension stripped).
            parent_folder_id: Optional folder to place the new Sheet in. Defaults
                to the same parent(s) as the source file.

        Returns:
            Dict with the new Sheet's metadata (id, name, mimeType, webViewLink)
            or error information.
        """
        try:
            if not source_file_id:
                return {"error": True, "message": "source_file_id cannot be empty"}

            # Validate the source is actually an .xlsx before attempting a
            # conversion that would otherwise produce a confusing API error.
            source = (
                self.service.files()
                .get(
                    fileId=source_file_id,
                    fields="id, name, mimeType, parents",
                    supportsAllDrives=True,
                )
                .execute()
            )
            source_mime = source.get("mimeType")
            if source_mime != XLSX_MIME_TYPE:
                return {
                    "error": True,
                    "error_type": "unsupported_type",
                    "message": (
                        f"Source file is '{source_mime}', not an .xlsx workbook. "
                        "Only .xlsx files can be converted with this tool."
                    ),
                    "operation": "convert_xlsx_to_google_sheet",
                }

            if not new_name:
                source_name = source.get("name", "Converted Sheet")
                new_name = (
                    source_name[: -len(".xlsx")]
                    if source_name.lower().endswith(".xlsx")
                    else source_name
                )

            body: dict[str, Any] = {
                "name": new_name,
                "mimeType": NATIVE_SHEET_MIME_TYPE,
            }
            if parent_folder_id:
                body["parents"] = [parent_folder_id]

            logger.info(
                f"Converting .xlsx '{source_file_id}' to native Google Sheet '{new_name}'."
            )
            new_sheet = (
                self.service.files()
                .copy(
                    fileId=source_file_id,
                    body=body,
                    fields="id, name, mimeType, webViewLink, parents",
                    supportsAllDrives=True,
                )
                .execute()
            )

            logger.info(
                f"Successfully converted to Google Sheet with ID: {new_sheet.get('id')}"
            )
            return new_sheet

        except HttpError as e:
            return self.handle_api_error("convert_xlsx_to_google_sheet", e)
        except Exception as e:
            logger.error(f"Non-API error in convert_xlsx_to_google_sheet: {str(e)}")
            return {
                "error": True,
                "error_type": "local_error",
                "message": f"Error converting .xlsx to Google Sheet: {str(e)}",
                "operation": "convert_xlsx_to_google_sheet",
            }

    def move_file(
        self,
        file_id: str,
        target_folder_id: str,
    ) -> dict[str, Any]:
        """
        Move a file into a target folder, removing its existing parents.

        Newly created native files (e.g. a Sheet made with
        sheets_create_spreadsheet) land in Drive root; this relocates them into
        the intended folder so deliverables are not stranded.

        Args:
            file_id: ID of the file to move.
            target_folder_id: ID of the destination folder.

        Returns:
            Dict with the file's id, name, and new parents, or error information.
        """
        try:
            if not file_id or not target_folder_id:
                return {
                    "error": True,
                    "message": "file_id and target_folder_id are required.",
                }

            # Look up current parents so they can be removed in the same call.
            current = (
                self.service.files()
                .get(fileId=file_id, fields="parents", supportsAllDrives=True)
                .execute()
            )
            previous_parents = ",".join(current.get("parents", []))

            logger.info(f"Moving file {file_id} to folder {target_folder_id}.")
            updated = (
                self.service.files()
                .update(
                    fileId=file_id,
                    addParents=target_folder_id,
                    removeParents=previous_parents,
                    fields="id, name, parents",
                    supportsAllDrives=True,
                )
                .execute()
            )
            return updated

        except HttpError as e:
            return self.handle_api_error("move_file", e)
        except Exception as e:
            logger.error(f"Non-API error in move_file: {str(e)}")
            return {
                "error": True,
                "error_type": "local_error",
                "message": f"Error moving file: {str(e)}",
                "operation": "move_file",
            }

    def rename_file(self, file_id: str, new_name: str) -> dict[str, Any]:
        """
        Rename an existing Drive file.

        Args:
            file_id: ID of the file to rename.
            new_name: The new name for the file.

        Returns:
            Dict with the file's id, name, and webViewLink, or error information.
        """
        try:
            if not file_id or not file_id.strip():
                return {"error": True, "message": "file_id is required."}
            if not new_name or not new_name.strip():
                return {"error": True, "message": "new_name cannot be empty."}

            logger.info(f"Renaming file {file_id} to '{new_name}'.")
            updated = (
                self.service.files()
                .update(
                    fileId=file_id,
                    body={"name": new_name.strip()},
                    fields="id, name, webViewLink",
                    supportsAllDrives=True,
                )
                .execute()
            )
            return updated

        except HttpError as e:
            return self.handle_api_error("rename_file", e)
        except Exception as e:
            logger.error(f"Non-API error in rename_file: {str(e)}")
            return {
                "error": True,
                "error_type": "local_error",
                "message": f"Error renaming file: {str(e)}",
                "operation": "rename_file",
            }

    def share_file(
        self, file_id: str, shared_drive_id: str | None = None
    ) -> dict[str, Any]:
        """Grant "anyone with the link → reader" on an existing file.

        For files that already exist (uploaded earlier, or created outside this
        tool) and need to become fetchable by their ``webContentLink`` without
        authentication. Idempotent — granting access to an already-shared file
        is a no-op. Returns the file's id + ``webContentLink`` plus the share
        outcome.

        Args:
            file_id: ID of the file to share.
            shared_drive_id: Optional Shared Drive id for ``supportsAllDrives``.

        Returns:
            Dict with id, webViewLink, webContentLink, shared, and (on failure)
            share_error; or error information.
        """
        try:
            if not file_id or not file_id.strip():
                return {"error": True, "message": "file_id is required."}

            logger.info(f"Sharing file {file_id} anyone-with-link → reader.")
            share = self._grant_anyone_reader(file_id, shared_drive_id)

            meta = (
                self.service.files()
                .get(
                    fileId=file_id,
                    fields="id, name, mimeType, webViewLink, webContentLink",
                    supportsAllDrives=True,
                )
                .execute()
            )
            return _with_download_url({**meta, **share})

        except HttpError as e:
            return self.handle_api_error("share_file", e)
        except Exception as e:
            logger.error(f"Non-API error in share_file: {str(e)}")
            return {
                "error": True,
                "error_type": "local_error",
                "message": f"Error sharing file: {str(e)}",
                "operation": "share_file",
            }

    def delete_file(self, file_id: str) -> dict[str, Any]:
        """
        Delete a file from Google Drive.

        Args:
            file_id: The ID of the file to delete

        Returns:
            Dict containing success status or error information
        """
        try:
            if not file_id:
                return {"success": False, "message": "File ID cannot be empty"}

            logger.info(f"Deleting file with ID: {file_id}")
            self.service.files().delete(fileId=file_id).execute()

            return {"success": True, "message": f"File {file_id} deleted successfully"}

        except Exception as e:
            return self.handle_api_error("delete_file", e)

    def list_shared_drives(self, page_size: int = 100) -> list[dict[str, Any]]:
        """
        Lists the user's shared drives.

        Args:
            page_size: Maximum number of shared drives to return. Max is 100.

        Returns:
            List of shared drive metadata dictionaries (id, name) or an error dictionary.
        """
        try:
            logger.info(f"Listing shared drives with page size: {page_size}")
            # API allows pageSize up to 100 for drives.list
            actual_page_size = min(max(1, page_size), 100)

            results = (
                self.service.drives()
                .list(pageSize=actual_page_size, fields="drives(id, name, kind)")
                .execute()
            )
            drives = results.get("drives", [])

            # Filter for kind='drive#drive' just to be sure, though API should only return these
            processed_drives = [
                {"id": d.get("id"), "name": d.get("name")}
                for d in drives
                if d.get("kind") == "drive#drive" and d.get("id") and d.get("name")
            ]
            logger.info(f"Found {len(processed_drives)} shared drives.")
            return processed_drives
        except HttpError as error:
            logger.error(f"Error listing shared drives: {error}")
            return self.handle_api_error("list_shared_drives", error)
        except Exception as e:
            logger.exception("Unexpected error listing shared drives")
            return {
                "error": True,
                "error_type": "unexpected_service_error",
                "message": str(e),
                "operation": "list_shared_drives",
            }
